import openpyxl
import matplotlib.pyplot as plt; plt.rcdefaults()
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image




wb = openpyxl.load_workbook('book.xlsx')
ws = wb.worksheets[0]
for i in range(2,ws.max_row):
    ws.cell(i,6).value = ws.cell(i,3).value + ws.cell(i,4).value + ws.cell(i,5).value
    ws.cell(i,7).value = str(int((ws.cell(i,3).value + ws.cell(i,4).value + ws.cell(i,5).value)/3)) + '%'

explode = (0, 0.1, 0, 0, 0)
objects = []
performance = []
val = []
for i in range(2,ws.max_row):
    objects.append(ws.cell(i,2).value)
    performance.append(ws.cell(i,6).value)
    val.append(ws.cell(i,9).value)

objects = tuple(objects)
y_pos = np.arange(len(objects))

plt.bar(y_pos, performance, align='center', alpha=0.5)
plt.xticks(y_pos, objects)
plt.ylabel('Max_Marks')
plt.title('Marks_Bar_Graph')


plt.savefig('image.png')
plt.savefig('pdff.pdf')
img = Image('image.png')
ws.add_image(img, 'A11')


fig1, ax1 = plt.subplots()
ax1.pie(val, explode=explode, labels=objects, autopct='%1.1f%%', shadow=True, startangle=90)
ax1.axis('equal')


plt.savefig('image1.png')
plt.savefig('pdff1.pdf')
img1 = Image('image1.png')
ws.add_image(img1, 'L11')

wb.save('book1.xlsx')

#print(objects)

#for i in ws.iter_rows(min_row = 1, max_row = ws.max_row, min_col = 1, max_col = 5):
    #print(i)
#    for j in i:
#        print(j.coordinate, end=" ")
#    print()





#ws.insert_row(0)
#ws.insert_column(0)
#ws.delete_rows(0)
#ws.delete_column(0)
